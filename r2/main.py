# onefile 
# Lib
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
from keras.models import Sequential
from keras.layers import Dense, BatchNormalization
from keras.optimizers import RMSprop
from keras.models import load_model
from sklearn.preprocessing import RobustScaler
from sklearn.metrics import mean_squared_error
from datetime import datetime
from io import StringIO
import tensorflow as tf
import openpyxl

# define the main function
def main():
    # Lib
    import pandas as pd
    import matplotlib.pyplot as plt
    import numpy as np
    import os
    from keras.models import Sequential
    from keras.layers import Dense, BatchNormalization
    from keras.optimizers import RMSprop
    from keras.models import load_model
    from sklearn.preprocessing import RobustScaler
    from sklearn.metrics import mean_squared_error
    from datetime import datetime
    from io import StringIO
    import tensorflow as tf
    import openpyxl
 

# Start the 2 loops
 
#   data to control looping
    ELstart = 50#130#20 # starting epochs value
    ELend = 550  # ending epochs value
    ELstep = 50 # step of epochs
    LayerLoopStart = 5 #9 #1 # starting layers value
    LayerLoopEnd = 31 # ending layers value
    LayerLoopStep = 5 # step of layers
    layersNum = LayerLoopStart # starting layers value
    # to collect all data in one excel sheet
    from datetime import datetime               
    fDateTime = datetime.now().strftime("%Y-%m-%d _ %H_%M_%S_")
    alloutputFile = 'out\\' +'out_All_' + fDateTime +' .xlsx'
    # add out path
    alloutputFile = os.path.join( os.getcwd(), alloutputFile)
 # end data 



    for j in range(LayerLoopStart,LayerLoopEnd,LayerLoopStep):
# start looping at layers
        # data to be changed at looping
        layersNum = j
        print('j : ',j)
        print('layersNum : ',layersNum)

        for i in range(ELstart,ELend,ELstep):

# start looping at epochs

# define the parameter
            # define the current working directory
            cwd = os.getcwd()
            # main input data 
            inputFolder = cwd
            readFolder = inputFolder
            print('os.listdir("readFolder") : ',os.listdir(readFolder))
            excel_files = [
                f for f in os.listdir(readFolder) if f.endswith(".xlsx") and not f.startswith("out")
            ]
            if not excel_files:
                raise ValueError("No Excel files found in the current directory")
            current_input_excel_file = excel_files[0]
            inputFile = os.path.join(inputFolder,current_input_excel_file)#'inputdata.xlsx')
            inputSheetName = "Sheet1"
# starting data will be changed at looping
            # data to be changed at looping
            myepochs = i
            print('i : ',i)
            print('myepochs : ',myepochs)
            #myepochs = 50
            mybatchSize = 32
            modelName = current_input_excel_file + '.layers.'+ str(layersNum) + '.epoch.' + str(myepochs)  +'.h5' #'testmodel.epoch.'+ str(myepochs) +'.h5'
            model = None    
            myOut = 'out\\' + 'num_layers_' +str(layersNum)+ '\\num_epochs_' + str(myepochs) + '_'
            outFolder = os.path.join(cwd,myOut)#cwd
            # calculated  parameters
            modelNamePath = os.path.join(inputFolder, modelName)      
            outputFile = 'out_' + modelName+' .xlsx'
            #from datetime import datetime               
            #fDateTime = datetime.now().strftime("%Y-%m-%d %H-%M-%S-")
            #alloutputFile = 'out_All_' + fDateTime +' .xlsx'
            summaryOutFile =  modelName + ' _ Summary .txt'
            outputSheetName = 'predicat_ '+ modelName+' '
            nameFigImg = 'fig_Sep_ '+ modelName+' .png'
            nameAllFigImg = 'fig_All_ '+ modelName+' .png'
            # add out path
            outputFile = os.path.join( outFolder, outputFile)
            #alloutputFile = os.path.join( cwd, alloutputFile)
            summaryOutFile = os.path.join( outFolder, summaryOutFile)
            nameFigImg = os.path.join( outFolder, nameFigImg)
            nameAllFigImg = os.path.join( outFolder, nameAllFigImg)
            print(' out folder : ',outFolder)
            print(' out modelName : ',modelName)
            print(' out nameAllFigImg : ',nameAllFigImg)
            
    # ending data will be changed at looping

# data will be calculated
            # data parameters    
            data = ''
            X = ''
            y = ''
            y_ln = ''
            X_train = ''  
            X_train_part = ''
            X_test = ''
            X_test_part = ''
            y_test_part = ''
            y_train_part = ''
            predictions = ''
            datap = pd.DataFrame(data=None)
            mergedData = ''
            outputpredicat = ''
            # Model parameteres     
            score = 0
            custom_loss_value = 0
            mse = 0
            rmse = 0
   
# Start the main function
            print("Starting main call")
# Read the data from the excel file in class of parameters
            data_all = pd.read_excel(inputFile, sheet_name=inputSheetName)

# at each distinct value of mass,s,N part remove the first three min value of Pt
            filtered_data = data_all.groupby(['mass', 's', 'N part']).apply(lambda x: x.nsmallest(3, 'Pt')).reset_index(drop=True)
            data_all = data_all.drop(filtered_data.index)

            datat = data_all .reset_index(drop=True)

            #temp_data_all = data_all  # data_all[data_all['spectrum']<60]
            datat = data_all .reset_index(drop=True)
            # make data sampled to randomized data as dr mohamed told us
            data = datat.sample(n=datat.count()[0]-1)
            data = data.reset_index(drop=True)
            # Split the data into input and output variables
            X = data[["mass", "s", "N part", "Pt"]]
            y = data["spectrum"].to_frame("spectrum")

            y_ln = np.log(y['spectrum'])
            #y_ln_df = pd.DataFrame(y_ln,columns=['ln spectrum'])

            print('y_ln -head : \n',y_ln.head)
            #print('y_ln_df -head : \n',y_ln_df)
            print('data_all : \n',data_all)
            print('x head : \n',X.head)
            print('y head : \n',y.head)
            print('data : \n',data)
            print ('data count : ',data.count())
            print('data size : ',data.size)    
            print('end get_data_in_Param')
            #value = input("Enter a value: ")

            """value = input("Enter a value: ")
            if value == "x":
                print("u will exit ...")
                exit()

            print("You entered:", value) """


# Normaliz input
            # Normalize the input
            from sklearn.preprocessing import RobustScaler
            # Create a RobustScaler object
            scaler = RobustScaler()
            # Fit the scaler to the input data and transform it
            X_normalized = scaler.fit_transform(X)
            # Print the normalized input data
            print('X_normalized')
            print(X_normalized)
            X_train = X_normalized
            print('end normaliz_data')        

# start looping at epochs

# Creat model : is commented when load model

            # Define the model
            model = Sequential(name= modelName)
            print('create_new_model : modelname \n', modelName)
            # Add the first dense layer
            model.add(Dense(40, input_dim=4, activation='relu'))

            # add the looped hidden layers
            for i in range(layersNum):
                model.add(Dense(40, activation='relu'))            

            # Add the hidden layers
            """  
            model.add(Dense(40, activation='relu'))
            model.add(Dense(40, activation='relu'))
            model.add(Dense(40, activation='relu'))
            model.add(Dense(80, activation='relu'))
            model.add(Dense(80, activation='relu'))
            model.add(Dense(40, activation='relu'))
            model.add(Dense(40, activation='relu'))
            model.add(Dense(40, activation='relu'))
            model.add(Dense(40, activation='relu'))
            """

            # Add the output layer
            model.add(Dense(1))
            ''' # compile the model      '''

            # define customized loss function 
            from keras import backend as BK
# define customized loss function with data as an argument
            def my_custom_loss(y_true, y_pred,data_for_lossfunction):
                y_pred = tf.cast(y_pred, tf.float64) # cast y_pred to float64
                y_true = tf.cast(y_true, tf.float64)
                # here write customized loss function
                diff = BK.square(y_true - y_pred)

                errorbar = BK.log( BK.abs(data_for_lossfunction['err2']))  + BK.log( BK.abs(data_for_lossfunction['err1']))
                #errorbar = np.log(errorbar)
                # Divide the squared difference by the error bars
                ratio = diff / BK.square(errorbar)
                # Sum the squared ratios
                loss = BK.sum(ratio)
                # divide on the number of data
                loss /= tf.cast(tf.shape(y_true)[0], tf.float64)
                # divide the loss by 2
                loss /= 2
                # clac the square root
                loss = BK.sqrt(loss)

                # return the loss : the root mean square error
                return loss

            # Compile the model with Levenberg-Marquardt optimizer
            optimizer = RMSprop(learning_rate=0.001, rho=0.001,)
            # use customized loss function   
            # Compile the model with the custom loss function and the data as an argument
            data_for_lossfunction = data
            print('dada_for_lossfunction : ', data_for_lossfunction)
            print('press key to continue')
            #  wait press key to continue
            #input()
            model.compile(loss=lambda y_true, y_pred: my_custom_loss(y_true, y_pred, data_for_lossfunction), optimizer='adam')    
            # model.compile(loss='mean_squared_error', optimizer=optimizer)
            print('after compiling model : model is: ', model)
            print('my epochs : ', myepochs)
            print('my batch size : ', mybatchSize)
            # fit on all data
            """  model.fit( X_train,  y, epochs= myepochs, 
                            batch_size= mybatchSize)  """
            # Set the validation split to 20%. and shuffel data
            from sklearn.model_selection import train_test_split
            ###
            X_train_part, X_test_part, y_train_part, y_test_part = train_test_split( X_train, y_ln,test_size=0.2,shuffle=True)

            """ 
             X_train_part =  X_train
             X_test_part =  X_train
             y_train_part =  y
             y_test_part =  y
            """
            model.fit( X_train_part,  y_train_part, epochs= myepochs, 
                            batch_size= mybatchSize, validation_split=0.2) 
            print('model after fitting : ', model)
            # Save the model
            modelNamePath = os.path.join( outFolder, modelName)    
            # Save the model 
            model.save( modelNamePath) 
            # model.save(os.path.join( outFolder, modelName)) 
            print('end create model')
            # Draw the model
            # from keras.utils.vis_utils import plot_model
            # Plot the model with custom arguments
            modelNameFig = modelNamePath+".png"
            from keras.utils import plot_model
            import pydot
            plot_model(model,to_file=modelNameFig,show_dtype=True,show_shapes=True,expand_nested=True,show_layer_activations=True,show_trainable=True)   
            #plot_model(model, to_file=modelNameFig, show_shapes=True, rankdir='LR')
            #tf.keras.utils.plot_model(model, to_file=modelNameFig, show_shapes=True)
            #value = input("Enter a value: ")
            print('end create model')
            #input('press key to continue')


    # Evalute the model
            # Evalute the model
            # Evaluate the model on the validation set
            evaluation = model.evaluate(X_test_part, y_test_part)
            print('evaluation : ',evaluation)

            # Get the value of the evaluation metric
            eval_metric_value = evaluation
            # Get the value of the customized loss function
            custom_loss_value = evaluation
            # Print the values
            score = eval_metric_value
            print("score " , score)
            print("Evaluation Metric Value:", eval_metric_value)
            print("Custom Loss Value:", custom_loss_value)

            #score = model.evaluate(X_test_part, y_test_part)   
            print('end evaluate')
            #input('press key to continue')

    # Predict the model

            # predict the model    
            from sklearn.metrics import mean_squared_error
            # Make predictions on new data
            X_test =pd.DataFrame(X_train) #scaler.transform(X)
            #X_test = scaler.fit_transform(X)
            print("new_data is - X_test : ")
            print(X_test)
            predictions = model.predict(X_test)
            predictions = predictions.flatten()
            predictions = pd.Series(predictions)
            predictions = predictions.to_frame('predictions')

            ###
            predictions = np.exp(predictions)
            print(" ### predictions is : ")
            print(predictions)
            #mse = mean_squared_error(y,predictions)
            mse = mean_squared_error(y,predictions)
            print('mse' , mse)       
            print('end predict')   
            #input('press key to continue')

    # Write output to excel 

            # merge date[] with predictions
            data['serr'] = data['err1']+data['err2']
            print('shape of datap',datap.shape)
            print('shape of predictions',predictions.shape)
            if(datap.empty):
                datap = pd.merge(data,predictions,left_index=True, right_index=True)
                print('shape of datap in if',datap.shape)
                pass
            else:
                pass
            print('shape of datap after if',datap.shape)
            # Write predictions , data to Excel file
# calculate rmse at excel sheet
            err1=data['err1'].to_frame('err1')
            err2=data['err2'].to_frame('err2')
            #SquareErrorForEachPoint = np.sqrt( ((datap['predictions']- datap['Spectrum'])/(err1- err2)))
            excelLnErrorbar =    (np.log(np.abs(err1['err1']))+ np.log(np.abs(err2['err2'])))
            # errorbar = np.log(errorbar)
            #excelLnErrorbar = np.log(excelLnErrorbar)
            SquareErrorForEachPoint =np.square( (datap['predictions']- datap['spectrum'])/(excelLnErrorbar))
            SquareErrorForEachPoint = pd.Series(SquareErrorForEachPoint)
            SquareErrorForEachPoint = SquareErrorForEachPoint.to_frame('SquareErrorForEachPoint')
            print('Square error for each point : ',SquareErrorForEachPoint)
            outputpredicat = pd.concat([datap, SquareErrorForEachPoint], axis=1)
            mysum =outputpredicat['SquareErrorForEachPoint'].sum()
            mycount =(outputpredicat['SquareErrorForEachPoint'].count()) -1
            rmse = mysum/mycount
            rmse /=2
            rmse = np.sqrt(rmse)
            #rmse = np.sqrt(mysum/mycount)
            #rmse = np.sqrt (np.average(outputpredicat['SquareErrorForEachPoint']))
            rmseX = pd.Series(rmse)
            rmseX = rmseX.to_frame('RMSE')
            rmseX['myepochs'] = myepochs
            rmseX['layersNum'] = layersNum
            rmseX['Score'] = score
            rmseX['mybatchSize'] = mybatchSize
            rmseX['custom_loss_value'] = custom_loss_value
            rmseX['modelName'] = modelName
            rmseX['current_input_excel_file'] = current_input_excel_file
            from datetime import datetime    
            # Get the current date and time
            current_datetime = datetime.now()    
            # Print the current date and time
            print("Current Date and Time:", current_datetime)
            formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")            
            print("Formatted Date and Time:", formatted_datetime)
            rmseX['currentDateTime'] = current_datetime
            rmseX['formatted_datetime'] = formatted_datetime

            #rmse = pd.DataFrame({'rmse': rmse})
            print('RMSE',rmseX)
            #custom_loss_value


            # Write the DataFrames to an Excel file with three sheets 
            with pd.ExcelWriter(outputFile) as writer:
                outputpredicat.to_excel(writer, sheet_name=outputSheetName, index=False)
                rmseX.to_excel(writer, sheet_name='RMSE', index=False)
            # append rmseX to fixed excel file
            #with pd.ExcelWriter(alloutputFile,mode='a') as writer:
            #    rmseX.to_excel(writer, sheet_name='RMSE', index=False)
           
            # append rmseX to fixed excel file
            # start here
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Check if the file exists
            if not os.path.isfile(alloutputFile):
                # Create a new Excel file and write the DataFrame
                rmseX.to_excel(alloutputFile, sheet_name='RMSE', index=False)
            else:
                # Append the DataFrame to the existing Excel file
                workbook = openpyxl.load_workbook(alloutputFile)
                
                # Select the 'RMSE' sheet or create it if it doesn't exist
                sheet_name = 'RMSE'
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(sheet_name)
                
                # Append the DataFrame to the sheet
                rows = dataframe_to_rows(rmseX, index=False, header=False)
                for row in rows:
                    sheet.append(row)
                
                # Save the changes to the Excel file
                workbook.save(alloutputFile)
           

            # end append to fixed excel sheet

           



            print('end write to excel ')
            #input('press key to continue')

    # Print model summery 

            from io import StringIO
            # summarize the model
            print('start summarize the model')
            with StringIO() as buf:
                model.summary(print_fn=lambda x: buf.write(x + '\n'))
                summary = buf.getvalue()
                summary += '\n score : ' + str(score)
                summary += '\n RMSE : ' + str(rmse)
                summary += '\n custom_loss_value : '+ str(custom_loss_value)
                summary += '\n formatted_datetime : '+ str(formatted_datetime)
                summary += '\n modelName : '+ str(modelName)
                summary += '\n mybatchSize : '+ str(mybatchSize)
                summary += '\n myepochs : '+ str(myepochs)
                #summary += '\n RMSE : ' + str(rmse.values)
            print('end summarize the model')       
            #with open(modelName +'-summary.txt', 'w') as f:
            with open(summaryOutFile,'w') as f:
                f.write(summary)
            print('start print summary')
            print('summary : ' , model.summary())
            print('model name : ',modelName)
            # print Root Mean Square Error that computed as Dr Mohamed want using err1&err2
            #print('RMSE',rmse['RMSE'].values)
            print('RMSE',rmse)
            # print("accuracy :" + str(accuracy))
            print("score " + str(score))
            print('myepochs '+ str(myepochs))
            #print("score " + str(score))
            print('end out model summary')
            #input('press key to continue')

    # Draw     
            # draw model_error_bar2
            # def draws_of_model_error_bar2:
            #  
            # for drawing in 2d i choose Pt as x-axis
            # draw after excel 
            # merge date[] with predictions
            if(datap.empty):
                datap = pd.merge(data,predictions,left_index=True, right_index=True)
                print('shape of datap',datap.shape)
                pass
            else:
                pass
            # datap : data + prediction
            #datap =pd.DataFrame(datap) # pd.DataFrame(datape12)# pd.DataFrame(datap)
            # xapf : datap after filteration
            xapf= datap.copy(deep=True)
            xapf['serr'] = xapf['err1']+ xapf['err2']
            print('xapf after serr', xapf)
            print('datap = ',datap)
            print('xapf : ',xapf)
            mass_filter = xapf['mass'].unique()[0]
            s_filter = xapf['s'].unique()[0]
            print('mass_filetr : ',mass_filter)
            print('s_filetr : ',s_filter)       
            #xapf = xapf[xapf['mass']== mass_filter] # 0.13957]   #0.493677] #0.13957]
            #xapf = xapf[xapf['s']== s_filter ] #7.7]
            #xapf = xapf[xapf['N part']==337]
            print('xapf modified : ',xapf)
            # detinct N Part values
            N_Part_Values  =xapf['N part'].unique() #xap['N part'].unique() # xapf['N part'].unique()
            print('Npart values : \n')
            for n in N_Part_Values:
                print('N is : ',n)
            print('Npart values : \n',N_Part_Values)    
            mergedData = xapf.copy()
            #yerror = mergedData['err2']
            #print('yerror : ',yerror.shape)
            #error = mergedData['err2'] #yerror.to_numpy()
            #print('error : ',error.shape)
            # Create a figure with  subplots
            #fig, axs = plt.subplots(nrows=len(N_Part_Values), ncols=1, figsize=(10, 100))
            for mass_item in xapf['mass'].unique():
                for s_item in xapf['s'].unique():
                    for n_part_item in xapf['N part'].unique():
                        print(' mass : ', mass_item)
                        print(' s : ',s_item)
                        print(' n part : ', n_part_item)
                        all_axis =    xapf.loc[(xapf['mass'] == mass_item) & (xapf['s'] == s_item) & (xapf['N part']==n_part_item)]
                        all_axis = all_axis.sort_values(by='Pt')
                        # filtered_df = df.loc[(df['age'] >= 25) & (df['age'] <= 40) & (df['gender'] == 'female')]
                        X_axis = all_axis['Pt']
                        #X_axis = xapf['Pt'][xapf['mass']==mass_item][xapf[xapf['s']== s_item]][xapf['N part']==n_part_item]
                        Yspectrum_axis = all_axis['spectrum']
                        #Yspectrum_axis = xapf['spectrum'][xapf['mass']==mass_item][xapf[xapf['s']== s_item]][xapf['N part']==n_part_item]
                        Ypredictions_axis = all_axis['predictions']
                        #Ypredictions_axis = xapf['predictions'][xapf['mass']==mass_item][xapf[xapf['s']== s_item]][xapf['N part']==n_part_item]
                        e_axis = all_axis['serr']
                        #e_axis = xapf['err2'][xapf['mass']==mass_item][xapf[xapf['s']== s_item]][xapf['N part']==n_part_item]
                        g_title = 'N_part = {:.0f}'.format(n_part_item) + ' ,mass = {:.6f}'.format(mass_item) + ' ,s = {:.1f}'.format(s_item)
                        gf_title = 'N_part = {:.0f}'.format(n_part_item) + ' ,mass = {:.6f}'.format(mass_item) + ' ,s = {:.1f}'.format(s_item)

                        #g_title += ' ' + ' score is  ' + '{:.3f}'.format(score)
                        #gf_title += '\n ' + ' score is  ' + '{:.3f}'.format(score)
                        if(datap.empty):
                            #g_title += ' ' + ' RMSE  is  ' + '{:.3f}'.format(rmse.values[1])
                            pass
                        else:
                            g_title += ' ' + ' RMSE  is  ' + '{:.3f}'.format(rmse)
                            gf_title += ' ' + ' RMSE  is  ' + '{:.3f}'.format(rmse)
                            pass
                        
                        g_title += ' myepochs '+ str(myepochs)
                        gf_title += ' '+' myepochs '+ str(myepochs)

                        if Yspectrum_axis.count() > 0:
                            #print('spectrum : ',Yspectrum_axis)

                            plt.errorbar(x=X_axis, 
                                 y=Yspectrum_axis, 
                                 yerr=e_axis, 
                                 fmt='o', color='blue',markersize=5,
                                 label=' spectrum ', #N_part = {}'.format(n_part_item),
                                 ecolor='green', elinewidth=3, capsize=1)
                            """ 
                            plt.scatter( x=X_axis, 
                                         y=Ypredictions_axis, 
                                         color='black', s=5,
                                         label=' predictions N_part = {}'.format(n))  
                            """
                            plt.plot( X_axis, 
                                         Ypredictions_axis, 
                                         color='red',
                                         label=' predictions ') # N_part = {}'.format(n_part_item)) 
                            max_s_p = int(max(max(Ypredictions_axis),max(Yspectrum_axis)))
                            y_scale = int(max(e_axis))                       
                            plt.xlim(left=0,right = int(max(X_axis)+0.1))
                            plt.ylim(bottom=0)
                            #plt.autoscale(enable=True, axis='x', tight=True)
                            plt.autoscale(enable=True, axis='y', tight=True)
                            # set the x-axis tick mark to be every 0.2 units
                            xticks=[i*0.2 for i in range(int(max(X_axis)/0.2)+1)]
                            plt.xticks(xticks)
                            plt.xlabel('Pt(GEV/C)')
                            plt.ylabel('Invariant Yield (GEV/C)^-2')
                            plt.title(gf_title)
                            plt.legend(loc='upper right')
                            #plt.legend(['Data'], loc='upper right')
                            #plt.legend(mergedData['spectrum'][mergedData['N part'] == n], loc='upper left')
                            plt.savefig(nameFigImg+'_'+ g_title +'_'+'.png')
                            plt.clf()
                            #plt.legend(['Data'], loc='upper left')
                            #plt.show()
                            pass
                        else:
                            pass 
            print("end plotting ",nameFigImg)
            print('end draws of model')
            #input("Press Enter to continue...")

    # summery data to show on screen only
            print('start print summary')
            print('start print summary')
            print('summary : ' , model.summary())
            print('model name : ',modelName)
            # print Root Mean Square Error that computed as Dr Mohamed want using err1&err2
            #print('RMSE',rmse['RMSE'].values)
            print('RMSE',rmse)
            # print("accuracy :" + str(accuracy))
            print("score " + str(score))
            print('myepochs '+ str(myepochs))
            #print("score " + str(score))
            print('end out model summary')
            #input("Press Enter to continue...")

# End loop for epochs
        print("End loop for epochs")
        #input("Press Enter to continue...")
# End loop for layers
    print("End loop for layers")



# End the main function
    print("The End of main ")

# call the main function
if __name__ == "__main__":
    print('main.py')
    main()
    print('main.py')

